import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

# Setting the Font Style of Cells.
italic24Font = Font(size=24, italic=True) # Create a font.
sheet['A1'].font = italic24Font # Apply the font to A1.
sheet['A1'] = 'Hello, world!'

# Font Objects.
fontObj1 = Font(name = 'Times New Roman', bold = True)
sheet['A1'].font = fontObj1

# Formulas.
sheet['A2'] = 200
sheet['A3'] = 300
sheet['A4'] = '=SUM(A2:A3)' # Set the formula.

# Setting Row Height and Column Width.
sheet['E1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70 # Set the height.
sheet.column_dimensions['B'].width = 20 # Set the width.

# Merging and Unmerging Cells.
sheet.merge_cells('A5:D8') # Merge all these cells.
sheet['A5'] = 'Twelve cells merged togethter'
sheet.merge_cells('C1:D1') # Merge these two cells.
sheet['C1'].font = Font(name='Times New Roman', size=18, bold =True)
sheet['C1'] = 'Two merged cells'
sheet.unmerge_cells('C1:D1') # Split these cells up

wb.save('styles.xlsx')
